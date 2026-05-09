"""FastAPI routes for running and downloading reconciliations."""

from __future__ import annotations

import hashlib
import io
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

from engine.reconciliation import reconcile_workbooks, serialise_result
from engine.normaliser import date_to_iso, decimal_to_float
from models.database import (
    get_bank_account,
    get_connection,
    get_match_report,
    get_run,
    insert_audit_log,
    insert_exception,
    insert_match,
    insert_run,
    insert_transaction,
    update_run,
    update_transaction_status,
)


router = APIRouter(prefix="/api/reconciliation", tags=["Reconciliation"])
BASE_DIR = Path(__file__).resolve().parents[1]
EXCEPTION_SLA_DAYS = {
    "unknown_dr": 1,
    "amount_mismatch": 1,
    "gib_unmatched": 3,
    "unknown_cr": 3,
    "stale_carry_forward": 0,
    "timing_difference": 0,
}


class ReconciliationRequest(BaseModel):
    """Payload for starting a new reconciliation run."""

    period_start: Optional[str] = None
    period_end: Optional[str] = None
    bank_statement_path: str
    bank_book_path: str
    previous_brs_path: Optional[str] = None
    previous_brs_sheet: Optional[str] = None
    bank_account_id: Optional[int] = None
    use_rag: bool = False


@router.post("/run")
async def start_reconciliation(req: ReconciliationRequest):
    """Run the deterministic reconciliation pipeline and persist a summary row."""

    if not os.path.exists(req.bank_statement_path):
        raise HTTPException(400, f"Bank statement file not found: {req.bank_statement_path}")
    if not os.path.exists(req.bank_book_path):
        raise HTTPException(400, f"Bank book file not found: {req.bank_book_path}")

    # Resolve bank account details
    bank_account = None
    if req.bank_account_id:
        with get_connection() as conn:
            bank_account = get_bank_account(conn, req.bank_account_id)
        if not bank_account:
            raise HTTPException(400, f"Bank account not found: {req.bank_account_id}")

    with get_connection() as conn:
        run_id = insert_run(
            conn,
            req.period_start or datetime.now().date().isoformat(),
            req.period_end or datetime.now().date().isoformat(),
            req.bank_statement_path,
            req.bank_book_path,
            req.previous_brs_path,
            bank_account_id=req.bank_account_id,
        )
        insert_audit_log(conn, "run_started", entity_type="run", entity_id=run_id)

    output_dir = BASE_DIR / "output"
    output_filename = f"BRS_run_{run_id}.xlsx"
    output_path = output_dir / output_filename

    try:
        result = reconcile_workbooks(
            statement_path=req.bank_statement_path,
            bank_book_path=req.bank_book_path,
            previous_brs_path=req.previous_brs_path,
            previous_brs_sheet=req.previous_brs_sheet,
            output_path=output_path,
            bank_account=bank_account,
            use_rag=req.use_rag,
        )
        summary = serialise_result(result)

        with get_connection() as conn:
            _persist_run_artifacts(conn, run_id, result)
            update_run(
                conn,
                run_id,
                status="completed",
                period_start=(req.period_start or result["statement"]["period_start"].isoformat()),
                period_end=(req.period_end or result["statement"]["period_end"].isoformat()),
                bank_book_balance=summary["totals"]["bank_book_balance"],
                bank_statement_balance=summary["totals"]["bank_statement_balance"],
                total_bank_stmt_entries=summary["statement_count"],
                total_bank_book_entries=summary["bank_book_count"],
                pass1_matches=result["matching"]["pass_counts"][1],
                pass2_matches=result["matching"]["pass_counts"][2],
                pass3_matches=result["matching"]["pass_counts"][3],
                pass4_matches=result["matching"]["pass_counts"][4],
                total_matched=summary["statement_count"] - len(result["matching"]["unmatched_statement"]),
                total_unmatched=len(result["matching"]["unmatched_statement"]) + len(result["matching"]["unmatched_book"]),
                total_pending=len(result["exceptions"]),
                brs_output_path=str(output_path),
                completed_at=datetime.now().isoformat(),
            )
            insert_audit_log(
                conn,
                "run_completed",
                entity_type="run",
                entity_id=run_id,
                details=summary,
            )
    except Exception as exc:  # pragma: no cover - surfaced to the API caller
        with get_connection() as conn:
            update_run(conn, run_id, status="failed")
            insert_audit_log(
                conn,
                "run_failed",
                entity_type="run",
                entity_id=run_id,
                details={"error": str(exc)},
            )
        raise HTTPException(500, str(exc)) from exc

    total_matched = summary["statement_count"] - len(result["matching"]["unmatched_statement"])
    total_unmatched = (
        len(result["matching"]["unmatched_statement"]) + len(result["matching"]["unmatched_book"])
    )
    auto_match_rate = round((total_matched / max(summary["statement_count"], 1)) * 100, 1)

    return {
        "run_id": run_id,
        "total_bank_stmt": summary["statement_count"],
        "total_bank_book": summary["bank_book_count"],
        "total_matched": total_matched,
        "total_unmatched": total_unmatched,
        "auto_match_rate": auto_match_rate,
        "carry_forward": len(result["matching"]["pending_carry_forward_items"]),
        "exception_count": len(result["exceptions"]),
        "pass_counts": result["matching"]["pass_counts"],
        **summary
    }


@router.get("/runs")
async def list_runs():
    """List previously created reconciliation runs."""

    with get_connection() as conn:
        rows = conn.execute("SELECT * FROM runs ORDER BY created_at DESC LIMIT 50").fetchall()
    return [dict(row) for row in rows]


@router.get("/run/{run_id}")
async def get_run_details(run_id: int):
    """Return the persisted summary for a previously completed run."""

    with get_connection() as conn:
        run = get_run(conn, run_id)
        if not run:
            raise HTTPException(404, "Run not found")

        # Reconstruct BRS section summary from stored transactions and carry-forward
        section_totals = _reconstruct_section_totals(conn, run_id)
        if section_totals and run.get("bank_book_balance") is not None:
            bb = run["bank_book_balance"]
            reconciled = (
                bb
                + section_totals["add_cheque_issued"]
                + section_totals["add_bank_credit"]
                - section_totals["less_cheque_deposit"]
                - section_totals["less_bank_debit"]
            )
            run["totals"] = {
                "bank_book_balance": bb,
                "bank_statement_balance": run.get("bank_statement_balance", 0),
                "add_cheque_issued": section_totals["add_cheque_issued"],
                "add_bank_credit": section_totals["add_bank_credit"],
                "less_cheque_deposit": section_totals["less_cheque_deposit"],
                "less_bank_debit": section_totals["less_bank_debit"],
                "reconciled_balance": reconciled,
                "difference": reconciled - (run.get("bank_statement_balance") or 0),
            }
            run["section_summary"] = section_totals["section_summary"]

    return run


def _reconstruct_section_totals(conn, run_id: int) -> dict | None:
    """Rebuild BRS section totals from persisted transactions and carry-forward rows."""

    # Unmatched book entries
    book_unmatched = conn.execute(
        "SELECT direction, SUM(amount) AS total, COUNT(*) AS cnt "
        "FROM transactions WHERE run_id=? AND source='bank_book' AND match_status='unmatched' "
        "GROUP BY direction",
        (run_id,),
    ).fetchall()
    # Unmatched statement entries
    stmt_unmatched = conn.execute(
        "SELECT direction, SUM(amount) AS total, COUNT(*) AS cnt "
        "FROM transactions WHERE run_id=? AND source='bank_statement' AND match_status='unmatched' "
        "GROUP BY direction",
        (run_id,),
    ).fetchall()
    # Carry-forward items by section
    cf_rows = conn.execute(
        "SELECT brs_section, SUM(amount) AS total, COUNT(*) AS cnt "
        "FROM carry_forward WHERE run_id=? GROUP BY brs_section",
        (run_id,),
    ).fetchall()

    if not book_unmatched and not stmt_unmatched and not cf_rows:
        return None

    sections = {
        "add_cheque_issued": 0.0,
        "add_bank_credit": 0.0,
        "less_cheque_deposit": 0.0,
        "less_bank_debit": 0.0,
    }
    counts = {s: 0 for s in sections}

    for row in book_unmatched:
        if row["direction"] == "OUT":
            sections["add_cheque_issued"] += row["total"] or 0
            counts["add_cheque_issued"] += row["cnt"] or 0
        else:
            sections["less_cheque_deposit"] += row["total"] or 0
            counts["less_cheque_deposit"] += row["cnt"] or 0

    for row in stmt_unmatched:
        if row["direction"] == "IN":
            sections["add_bank_credit"] += row["total"] or 0
            counts["add_bank_credit"] += row["cnt"] or 0
        else:
            sections["less_bank_debit"] += row["total"] or 0
            counts["less_bank_debit"] += row["cnt"] or 0

    for row in cf_rows:
        sec = row["brs_section"]
        if sec in sections:
            sections[sec] += row["total"] or 0
            counts[sec] += row["cnt"] or 0

    return {
        **sections,
        "section_summary": {
            s: {"count": counts[s], "total": sections[s]} for s in sections
        },
    }


@router.get("/run/{run_id}/matches")
async def get_run_matches(run_id: int):
    """Return the persisted matched report for a reconciliation run."""

    with get_connection() as conn:
        report = get_match_report(conn, run_id)
        if not report:
            raise HTTPException(404, "Run not found")
    return report


@router.get("/run/{run_id}/matches/download")
async def download_matches_excel(run_id: int):
    """Download the matched report as an Excel workbook."""

    with get_connection() as conn:
        report = get_match_report(conn, run_id)
        if not report:
            raise HTTPException(404, "Run not found")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── Summary sheet ────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws_summary.append(["Matched Report — Run #" + str(run_id)])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([])
    ws_summary.append(["Period", f"{report.get('period_start', '')} to {report.get('period_end', '')}"])
    ws_summary.append(["Completed At", report.get("completed_at", "")])
    ws_summary.append(["Total Match Groups", report["match_count"]])
    ws_summary.append(["Statement Entries Matched", report["statement_entry_count"]])
    ws_summary.append(["Book Entries Matched", report["bank_book_entry_count"]])
    ws_summary.append(["Total Matched Amount", report["total_matched_amount"]])
    ws_summary.append([])

    # Pass-wise breakdown
    from collections import Counter
    pass_counts = Counter()
    type_counts = Counter()
    for m in report["matches"]:
        pass_counts[m["pass_number"]] += 1
        type_counts[m["match_type"]] += 1

    ws_summary.append(["Pass-wise Breakdown"])
    ws_summary[f"A{ws_summary.max_row}"].font = header_font
    for pn in sorted(pass_counts):
        ws_summary.append([f"  Pass {pn}", pass_counts[pn]])

    ws_summary.append([])
    ws_summary.append(["Match Type Breakdown"])
    ws_summary[f"A{ws_summary.max_row}"].font = header_font
    for mt, cnt in type_counts.most_common():
        ws_summary.append([f"  {mt}", cnt])

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    # ── Matches detail sheet ─────────────────────────────────────────
    ws = wb.create_sheet("Matches")

    detail_headers = [
        "Match #", "Pass", "Match Type", "Confidence", "Matched Amount",
        "Notes",
        "Stmt Date", "Stmt Amount", "Stmt Direction", "Stmt Description",
        "Stmt Refs",
        "Book Date", "Book Amount", "Book Direction", "Book Narration",
        "Book Voucher Type", "Book Voucher No", "Book Refs",
    ]
    ws.append(detail_headers)
    for col_idx, _ in enumerate(detail_headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for idx, match in enumerate(report["matches"], 1):
        stmt_entries = match.get("statement_entries", [])
        book_entries = match.get("bank_book_entries", [])
        max_rows = max(len(stmt_entries), len(book_entries), 1)

        for row_offset in range(max_rows):
            row_data = []
            # Common fields — only on first row of each group
            if row_offset == 0:
                row_data = [
                    idx,
                    match["pass_number"],
                    match["match_type"],
                    match["confidence"],
                    match["matched_amount"],
                    match["notes"],
                ]
            else:
                row_data = ["", "", "", "", "", ""]

            # Statement entry
            if row_offset < len(stmt_entries):
                s = stmt_entries[row_offset]
                row_data += [
                    s.get("transaction_date", ""),
                    s.get("amount", ""),
                    s.get("direction", ""),
                    s.get("description") or s.get("narration", ""),
                    ", ".join(s.get("references", [])),
                ]
            else:
                row_data += ["", "", "", "", ""]

            # Book entry
            if row_offset < len(book_entries):
                b = book_entries[row_offset]
                row_data += [
                    b.get("transaction_date", ""),
                    b.get("amount", ""),
                    b.get("direction", ""),
                    b.get("narration") or b.get("description", ""),
                    b.get("voucher_type", ""),
                    b.get("voucher_no", ""),
                    ", ".join(b.get("references", [])),
                ]
            else:
                row_data += ["", "", "", "", "", "", ""]

            ws.append(row_data)

    # Auto-size columns
    col_widths = [8, 6, 22, 10, 15, 30, 12, 14, 6, 45, 20, 12, 14, 6, 45, 12, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Write to bytes ───────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Matched_Report_Run_{run_id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/run/{run_id}/download")
async def download_brs(run_id: int):
    """Download the generated BRS workbook for a run."""

    with get_connection() as conn:
        run = get_run(conn, run_id)
        if not run:
            raise HTTPException(404, "Run not found")

        output_path = run.get("brs_output_path")
        if not output_path or not os.path.exists(output_path):
            raise HTTPException(404, "BRS output file not found")

    return FileResponse(
        output_path,
        filename=os.path.basename(output_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _persist_run_artifacts(conn, run_id: int, result: dict):
    """Persist transactions, matches, exceptions, and carry-forward rows for a run."""

    statement_rows = result["statement"]["transactions"]
    bank_book_rows = result["bank_book"]["transactions"]
    carry_forward_items = result["matching"]["pending_carry_forward_items"]
    matches = result["matching"]["matches"]

    statement_pass_map, bank_book_pass_map = _build_match_pass_maps(matches)
    statement_ids = {}
    bank_book_ids = {}
    carry_forward_ids = {}

    for row in statement_rows:
        transaction_id = insert_transaction(
            conn,
            run_id,
            "bank_statement",
            date_to_iso(row["value_date"]) or "",
            decimal_to_float(row["amount"]),
            row["direction"],
            references=row.get("refs", []),
            description=row["description"],
            cheque_no=row.get("cheque_no"),
            transaction_id=row.get("transaction_id"),
            original_row=row["row_number"],
            sha256_hash=row["row_hash"],
        )
        statement_ids[row["row_number"]] = transaction_id
        _sync_transaction_status(
            conn,
            transaction_id,
            matched=row.get("matched", False),
            pass_number=statement_pass_map.get(row["row_number"]),
        )

    for row in bank_book_rows:
        transaction_id = insert_transaction(
            conn,
            run_id,
            "bank_book",
            date_to_iso(row["voucher_date"]) or "",
            decimal_to_float(row["amount"]),
            row["direction"],
            references=row.get("refs", []),
            narration=row["narration"],
            description=row["particulars"],
            voucher_type=row.get("voucher_type"),
            voucher_no=row.get("voucher_no"),
            cheque_no=row.get("cheque_no"),
            original_row=row["row_number"],
            sha256_hash=row["row_hash"],
        )
        bank_book_ids[row["row_number"]] = transaction_id
        _sync_transaction_status(
            conn,
            transaction_id,
            matched=row.get("matched", False),
            pass_number=bank_book_pass_map.get(row["row_number"]),
        )

    for item in carry_forward_items:
        transaction_id = insert_transaction(
            conn,
            run_id,
            "carry_forward",
            date_to_iso(item["original_date"]) or "",
            decimal_to_float(item["amount"]),
            item["direction"],
            references=item.get("refs", []),
            narration=item["remarks"],
            description=item["remarks"],
            cheque_no=item.get("cheque_no"),
            original_row=item["row_number"],
            sha256_hash=_carry_forward_hash(run_id, item),
        )
        carry_forward_ids[item["row_number"]] = transaction_id
        conn.execute(
            """INSERT INTO carry_forward
               (run_id, brs_section, original_date, remarks, cheque_no, amount, cleared_date, source_run_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                run_id,
                item["section"],
                date_to_iso(item["original_date"]) or "",
                item["remarks"],
                item.get("cheque_no"),
                decimal_to_float(item["amount"]),
                date_to_iso(item.get("cleared_on")),
                None,
            ),
        )

    for match in matches:
        pass_number = match.get("pass_number")
        if pass_number not in {1, 2, 3, 4}:
            continue

        insert_match(
            conn,
            run_id,
            pass_number,
            match["match_type"],
            [statement_ids[row_number] for row_number in match.get("statement_rows", []) if row_number in statement_ids],
            [bank_book_ids[row_number] for row_number in match.get("book_rows", []) if row_number in bank_book_ids],
            decimal_to_float(match["amount"]),
            notes=match.get("notes"),
        )

    for exception in result["exceptions"]:
        transaction_id = _exception_transaction_id(
            exception,
            statement_ids=statement_ids,
            bank_book_ids=bank_book_ids,
            carry_forward_ids=carry_forward_ids,
        )
        if not transaction_id:
            continue

        exception_id = insert_exception(
            conn,
            run_id,
            transaction_id,
            exception["exception_type"],
            exception["brs_section"],
            sla_days=EXCEPTION_SLA_DAYS.get(exception["exception_type"], 3),
        )
        insert_audit_log(
            conn,
            "exception_created",
            entity_type="exception",
            entity_id=exception_id,
            details={
                "exception_type": exception["exception_type"],
                "source": exception["source"],
                "row_number": exception["row_number"],
            },
        )


def _build_match_pass_maps(matches: list[dict]) -> tuple[dict[int, int], dict[int, int]]:
    """Build row-number-to-pass lookups from the in-memory match results."""

    statement_pass_map: dict[int, int] = {}
    bank_book_pass_map: dict[int, int] = {}
    for match in matches:
        pass_number = match.get("pass_number")
        if pass_number not in {1, 2, 3, 4}:
            continue
        for row_number in match.get("statement_rows", []):
            statement_pass_map[row_number] = pass_number
        for row_number in match.get("book_rows", []):
            bank_book_pass_map[row_number] = pass_number
    return statement_pass_map, bank_book_pass_map


def _sync_transaction_status(conn, transaction_id: int, *, matched: bool, pass_number: int | None):
    """Persist the transaction match state without losing non-pass matches."""

    if matched and pass_number:
        update_transaction_status(conn, transaction_id, "matched", pass_number)
        return
    if matched:
        update_transaction_status(conn, transaction_id, "matched")


def _carry_forward_hash(run_id: int, item: dict) -> str:
    """Return a deterministic hash for persisted carry-forward rows."""

    raw = "|".join(
        [
            str(run_id),
            item["section"],
            date_to_iso(item["original_date"]) or "",
            str(item["amount"]),
            item["remarks"],
        ]
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _exception_transaction_id(
    exception: dict,
    *,
    statement_ids: dict[int, int],
    bank_book_ids: dict[int, int],
    carry_forward_ids: dict[int, int],
) -> int | None:
    """Resolve the persisted transaction ID for an exception payload."""

    if exception["source"] == "bank_statement":
        return statement_ids.get(exception["row_number"])
    if exception["source"] == "bank_book":
        return bank_book_ids.get(exception["row_number"])
    return carry_forward_ids.get(exception["row_number"])
