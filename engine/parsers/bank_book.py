"""Parser for the ERP bank book ledger workbook."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any
import re

import openpyxl

from engine.normaliser import (
    amounts_equal,
    date_to_iso,
    hash_book_row,
    normalise_amount,
    normalise_date,
    normalise_direction_book,
)
from engine.reference_extractor import extract_tn_nos, extract_transaction_date


SHEET_PATTERN = re.compile(r"^BANK BOOK LEDGER", re.IGNORECASE)
REQUIRED_COLUMNS = [
    "Branch",
    "Voucher Date",
    "Voucher No",
    "Particluars",
    "Voucher Type",
    "Debit",
    "Credit",
    "Cheque No",
    "Narration",
]
OPTIONAL_COLUMNS = ["Voucher Type Name", "Invoice No."]
VALID_VOUCHER_TYPES = {"REC", "PMT", "CNT"}

# Map full-length voucher type names to the canonical abbreviations used
# throughout the matching engine.
VOUCHER_TYPE_ALIASES: dict[str, str] = {
    "RECEIPT": "REC",
    "PAYMENT": "PMT",
    "CONTRA": "CNT",
    "REC": "REC",
    "PMT": "PMT",
    "CNT": "CNT",
}


def parse_bank_book(
    filepath: str | Path,
    *,
    account_id: str = "082401002764",
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Parse the BWU ERP bank book and compute the month-end closing balance."""

    workbook = openpyxl.load_workbook(filepath, data_only=True)
    worksheet = _select_sheet(workbook, sheet_name)

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Bank book sheet is empty.")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    column_index = {header: idx for idx, header in enumerate(headers)}
    missing = [column for column in REQUIRED_COLUMNS if column not in column_index]
    if missing:
        raise ValueError(
            f"Bank book schema mismatch. Missing columns: {missing}. Found: {headers}"
        )

    transactions: list[dict[str, Any]] = []
    opening_balance = Decimal("0.00")

    for excel_row_number, row in enumerate(rows[1:], start=2):
        voucher_date = normalise_date(row[column_index["Voucher Date"]])
        particulars = str(row[column_index["Particluars"]] or "").strip()
        raw_voucher_type = str(row[column_index["Voucher Type"]] or "").strip().upper()
        voucher_type = VOUCHER_TYPE_ALIASES.get(raw_voucher_type, raw_voucher_type)

        if particulars == "Opening Balance":
            opening_balance = normalise_amount(row[column_index["Debit"]])
            continue
        if voucher_date is None or voucher_type not in VALID_VOUCHER_TYPES:
            continue

        debit = normalise_amount(row[column_index["Debit"]])
        credit = normalise_amount(row[column_index["Credit"]])
        direction = normalise_direction_book(debit, credit)
        amount = debit if direction == "IN" else credit
        narration = str(row[column_index["Narration"]] or "").strip()
        ledger_text = " ".join(filter(None, [narration, particulars]))

        record = {
            "kind": "book",
            "row_number": excel_row_number,
            "branch": _clean_optional_text(row[column_index["Branch"]]),
            "voucher_date": voucher_date,
            "voucher_no": _clean_optional_text(row[column_index["Voucher No"]]),
            "particulars": particulars,
            "voucher_type": voucher_type,
            "voucher_type_name": _clean_optional_text(row[column_index["Voucher Type Name"]]) if "Voucher Type Name" in column_index else None,
            "invoice_no": _clean_optional_text(row[column_index["Invoice No."]]) if "Invoice No." in column_index else None,
            "debit": debit,
            "credit": credit,
            "cheque_no": _clean_optional_text(row[column_index["Cheque No"]]),
            "narration": narration,
            "direction": direction,
            "amount": amount,
            "refs": extract_tn_nos(ledger_text),
            "txn_date": extract_transaction_date(ledger_text),
            "match_state": "unmatched",
            "matched": False,
        }
        record["row_hash"] = hash_book_row(
            account_id,
            date_to_iso(voucher_date) or "",
            str(record["voucher_no"] or ""),
            str(amount),
            direction,
        )
        transactions.append(record)

    closing_balance = opening_balance
    for record in transactions:
        closing_balance += record["debit"]
        closing_balance -= record["credit"]

    summary_closing_balance = _extract_summary_closing_balance(rows, column_index)
    if summary_closing_balance is not None and not amounts_equal(summary_closing_balance, closing_balance):
        raise ValueError(
            "Computed bank book closing balance does not match the workbook summary row."
        )

    workbook.close()
    return {
        "sheet_name": worksheet.title,
        "transactions": transactions,
        "count": len(transactions),
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "period_start": transactions[0]["voucher_date"] if transactions else None,
        "period_end": transactions[-1]["voucher_date"] if transactions else None,
    }


def _select_sheet(workbook: openpyxl.Workbook, explicit_name: str | None) -> openpyxl.worksheet.worksheet.Worksheet:
    """Select the ledger sheet by explicit name or expected naming pattern."""

    if explicit_name:
        if explicit_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{explicit_name}' not found in {workbook.sheetnames}")
        return workbook[explicit_name]

    for candidate in workbook.sheetnames:
        if SHEET_PATTERN.search(candidate):
            return workbook[candidate]

    for candidate in workbook.worksheets:
        if _sheet_has_bank_book_schema(candidate):
            return candidate

    if len(workbook.sheetnames) == 1:
        return workbook.active

    raise ValueError(f"No bank book sheet found in workbook: {workbook.sheetnames}")


def _sheet_has_bank_book_schema(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
    """Detect the ledger sheet from row 1 headers even when the tab is named ``Sheet1``."""

    header_row = [cell.value for cell in worksheet[1]]
    headers = {str(value).strip() for value in header_row if value is not None}
    return all(column in headers for column in REQUIRED_COLUMNS)


def _extract_summary_closing_balance(
    rows: list[tuple[Any, ...]],
    column_index: dict[str, int],
) -> Decimal | None:
    """Read the optional ``Closing Balance`` summary row at the bottom of the sheet."""

    for row in rows[1:]:
        if str(row[0] or "").strip() == "Closing Balance":
            return normalise_amount(row[column_index["Credit"]])
    return None


def _clean_optional_text(raw: Any) -> str | None:
    """Return a cleaned text value or ``None`` for blank cells."""

    text = str(raw).strip() if raw is not None else ""
    return text or None
