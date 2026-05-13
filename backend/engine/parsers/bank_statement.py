"""Parser for bank statement workbooks.

Supported bank statement formats:
- ICICI Bank: No., Transaction ID, Value Date, Txn Posted Date, ChequeNo., 
              Description, Cr/Dr, Transaction Amount(INR), Available Balance(INR)
- Ujjivan SFB: Date, Particular, Chq./Ref.no., Withdrawal, Deposit, Balance Amount
- Bandhan Bank (HDFC-like): Date, Description, Value Date, Reference No., 
                            Cheque No./Instrument No., Credit, Debit, Balance
- HDFC Bank: Date, Narration, Chq./Ref.No., Value Dt, Withdrawal Amt., 
             Deposit Amt., Closing Balance
- Axis Bank: S. No., Transaction Date, Value Date, Cheque Number, 
             Transaction Particulars, Amount, Transaction Type, Balance
- IDBI Bank: GL. Date, Value Date, Tran Id, Instrmnt Number, Particulars, 
             Transaction Debit Amount, Transaction Credit Amount, Balance
- Karur Vysya Bank: Transaction Date, Value Date, Branch, Cheque No., 
                    Description, Debit, Credit, Balance
- SBI Bank: Txn Date, Value Date, Description, Ref No./Cheque No., 
            Branch Code, Debit, Credit, Balance
"""

from __future__ import annotations

from pathlib import Path
from typing import Any
import re

import openpyxl

from engine.normaliser import (
    date_to_iso,
    hash_bank_stmt_row,
    normalise_amount,
    normalise_date,
    normalise_direction_stmt,
)
from engine.reference_extractor import extract_ref_from_description


STATEMENT_HEADER_SCAN_LIMIT = 15
# Ujjivan files may have extensive account metadata before the header row.
UJJIVAN_HEADER_SCAN_LIMIT = 50
SHEET_PATTERN = re.compile(r"^BANK STATEMENT", re.IGNORECASE)

# ---------------------------------------------------------------------------
# ICICI column schema
# ---------------------------------------------------------------------------
ICICI_REQUIRED_COLUMNS = [
    "No.",
    "Transaction ID",
    "Value Date",
    "Txn Posted Date",
    "ChequeNo.",
    "Description",
    "Cr/Dr",
    "Transaction Amount(INR)",
    "Available Balance(INR)",
]
# Keep the old name as an alias so existing tests still pass.
REQUIRED_COLUMNS = ICICI_REQUIRED_COLUMNS

# ---------------------------------------------------------------------------
# Ujjivan SFB column schema
# ---------------------------------------------------------------------------
UJJIVAN_REQUIRED_COLUMNS = [
    "Date",
    "Particular",
    "Chq./Ref.no.",
    "Withdrawal",
    "Deposit",
    "Balance Amount",
]

# ---------------------------------------------------------------------------
# HDFC column schema (Bandhan Bank also uses this format)
# ---------------------------------------------------------------------------
HDFC_REQUIRED_COLUMNS = [
    "Date",
    "Description",
    "Value Date",
    "Reference No.",
    "Credit",
    "Debit",
    "Balance",
]
# HDFC files have extensive metadata headers before data rows
HDFC_HEADER_SCAN_LIMIT = 20

# ---------------------------------------------------------------------------
# HDFC Bank (actual HDFC format) column schema
# ---------------------------------------------------------------------------
HDFC_ACTUAL_REQUIRED_COLUMNS = [
    "Date",
    "Narration",
    "Chq./Ref.No.",
    "Value Dt",
    "Withdrawal Amt.",
    "Deposit Amt.",
    "Closing Balance",
]
HDFC_ACTUAL_HEADER_SCAN_LIMIT = 30

# ---------------------------------------------------------------------------
# Axis Bank column schema
# ---------------------------------------------------------------------------
AXIS_REQUIRED_COLUMNS = [
    "S. No.",
    "Transaction Date",
    "Value Date",
    "Transaction Particulars",
    "Amount",
    "Transaction Type",
    "Balance",
]
AXIS_HEADER_SCAN_LIMIT = 20

# ---------------------------------------------------------------------------
# IDBI Bank column schema
# ---------------------------------------------------------------------------
IDBI_REQUIRED_COLUMNS = [
    "GL.",
    "Value Date",
    "Tran Id",
    "Particulars",
]
IDBI_HEADER_SCAN_LIMIT = 30

# ---------------------------------------------------------------------------
# Karur Vysya (KV) Bank column schema
# ---------------------------------------------------------------------------
KV_REQUIRED_COLUMNS = [
    "Transaction Date",
    "Value Date",
    "Description",
    "Debit",
    "Credit",
    "Balance",
]
KV_HEADER_SCAN_LIMIT = 20

# ---------------------------------------------------------------------------
# SBI Bank column schema
# ---------------------------------------------------------------------------
SBI_REQUIRED_COLUMNS = [
    "Txn Date",
    "Value Date",
    "Description",
    "Debit",
    "Credit",
    "Balance",
]
SBI_HEADER_SCAN_LIMIT = 25


def parse_bank_statement(
    filepath: str | Path,
    *,
    account_id: str = "082401002764",
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Auto-detect the bank format and parse into normalised transaction dicts."""

    workbook = openpyxl.load_workbook(filepath, data_only=True)
    worksheet = _select_sheet(workbook, sheet_name)

    rows = list(worksheet.iter_rows(values_only=True))

    # --- Try ICICI format first ---
    header_row_index = _find_header_row(rows, ICICI_REQUIRED_COLUMNS, STATEMENT_HEADER_SCAN_LIMIT)
    if header_row_index is not None:
        result = _parse_icici(rows, header_row_index, account_id, worksheet.title)
        workbook.close()
        return result

    # --- Try Ujjivan SFB format ---
    header_row_index = _find_header_row(rows, UJJIVAN_REQUIRED_COLUMNS, UJJIVAN_HEADER_SCAN_LIMIT)
    if header_row_index is not None:
        result = _parse_ujjivan(rows, header_row_index, account_id, worksheet.title)
        workbook.close()
        return result

    # --- Try HDFC/Bandhan format (Credit/Debit columns with Reference No.) ---
    header_row_index = _find_header_row(rows, HDFC_REQUIRED_COLUMNS, HDFC_HEADER_SCAN_LIMIT)
    if header_row_index is not None:
        result = _parse_hdfc(rows, header_row_index, account_id, worksheet.title)
        workbook.close()
        return result

    # --- Try actual HDFC Bank format (Withdrawal Amt./Deposit Amt.) ---
    header_row_index = _find_header_row(rows, HDFC_ACTUAL_REQUIRED_COLUMNS, HDFC_ACTUAL_HEADER_SCAN_LIMIT)
    if header_row_index is not None:
        result = _parse_hdfc_actual(rows, header_row_index, account_id, worksheet.title)
        workbook.close()
        return result

    # --- Try Axis Bank format ---
    header_row_index = _find_header_row(rows, AXIS_REQUIRED_COLUMNS, AXIS_HEADER_SCAN_LIMIT)
    if header_row_index is not None:
        result = _parse_axis(rows, header_row_index, account_id, worksheet.title)
        workbook.close()
        return result

    # --- Try IDBI Bank format ---
    header_row_index = _find_header_row(rows, IDBI_REQUIRED_COLUMNS, IDBI_HEADER_SCAN_LIMIT)
    if header_row_index is not None:
        result = _parse_idbi(rows, header_row_index, account_id, worksheet.title)
        workbook.close()
        return result

    # --- Try Karur Vysya (KV) Bank format ---
    header_row_index = _find_header_row(rows, KV_REQUIRED_COLUMNS, KV_HEADER_SCAN_LIMIT)
    if header_row_index is not None:
        result = _parse_kv(rows, header_row_index, account_id, worksheet.title)
        workbook.close()
        return result

    # --- Try SBI Bank format ---
    header_row_index = _find_header_row(rows, SBI_REQUIRED_COLUMNS, SBI_HEADER_SCAN_LIMIT)
    if header_row_index is not None:
        result = _parse_sbi(rows, header_row_index, account_id, worksheet.title)
        workbook.close()
        return result

    workbook.close()
    raise ValueError(
        "Bank statement sheet does not match any supported format "
        "(ICICI, Ujjivan SFB, HDFC, Bandhan, Axis, IDBI, Karur Vysya, or SBI)."
    )


# ===================================================================
# ICICI parsing
# ===================================================================

def _parse_icici(
    rows: list[tuple[Any, ...]],
    header_row_index: int,
    account_id: str,
    sheet_title: str,
) -> dict[str, Any]:
    """Parse ICICI-format statement rows."""

    data_start = header_row_index + 1
    headers = [str(v).strip() if v is not None else "" for v in rows[header_row_index]]
    col = {h: i for i, h in enumerate(headers)}

    transactions: list[dict[str, Any]] = []
    closing_balance = normalise_amount(0)

    for excel_row, row in enumerate(rows[data_start:], start=data_start + 1):
        if row[col["No."]] is None:
            continue

        value_date = normalise_date(row[col["Value Date"]])
        description = str(row[col["Description"]] or "").strip()
        direction = normalise_direction_stmt(row[col["Cr/Dr"]])
        amount = normalise_amount(row[col["Transaction Amount(INR)"]])
        available_balance = normalise_amount(row[col["Available Balance(INR)"]])
        transaction_id = str(row[col["Transaction ID"]] or "").strip()

        record = {
            "kind": "statement",
            "row_number": excel_row,
            "statement_no": row[col["No."]],
            "transaction_id": transaction_id,
            "value_date": value_date,
            "posted_date": normalise_date(row[col["Txn Posted Date"]]),
            "cheque_no": _clean_optional_text(row[col["ChequeNo."]]),
            "description": description,
            "direction": direction,
            "amount": amount,
            "available_balance": available_balance,
            "refs": extract_ref_from_description(description),
            "is_fd": transaction_id.startswith("M"),
            "match_state": "unmatched",
            "matched": False,
        }
        record["row_hash"] = hash_bank_stmt_row(
            account_id,
            date_to_iso(value_date) or "",
            str(amount),
            direction,
            description,
        )
        transactions.append(record)
        closing_balance = available_balance

    return {
        "sheet_name": sheet_title,
        "transactions": transactions,
        "count": len(transactions),
        "closing_balance": closing_balance,
        "period_start": transactions[0]["value_date"] if transactions else None,
        "period_end": transactions[-1]["value_date"] if transactions else None,
    }


# ===================================================================
# Ujjivan SFB parsing
# ===================================================================

def _parse_ujjivan(
    rows: list[tuple[Any, ...]],
    header_row_index: int,
    account_id: str,
    sheet_title: str,
) -> dict[str, Any]:
    """Parse Ujjivan SFB-format statement rows.

    Ujjivan uses separate Withdrawal / Deposit columns instead of Cr/Dr,
    has no Transaction ID or statement number, and dates may be datetime
    objects or ``"DD-MM-YY"`` strings.
    """

    data_start = header_row_index + 1
    headers = [str(v).strip() if v is not None else "" for v in rows[header_row_index]]
    col = {h: i for i, h in enumerate(headers)}

    transactions: list[dict[str, Any]] = []
    closing_balance = normalise_amount(0)
    seq = 0

    for excel_row, row in enumerate(rows[data_start:], start=data_start + 1):
        # Skip rows without a parseable date (footer / summary rows).
        raw_date = row[col["Date"]] if col["Date"] < len(row) else None
        value_date = normalise_date(raw_date)
        if value_date is None:
            continue

        description = str(row[col["Particular"]] or "").strip()
        if not description:
            continue

        withdrawal = normalise_amount(row[col["Withdrawal"]] if col["Withdrawal"] < len(row) else 0)
        deposit = normalise_amount(row[col["Deposit"]] if col["Deposit"] < len(row) else 0)

        if withdrawal > 0:
            direction = "OUT"
            amount = withdrawal
        elif deposit > 0:
            direction = "IN"
            amount = deposit
        else:
            continue  # skip zero-amount rows

        bal_idx = col["Balance Amount"]
        available_balance = normalise_amount(row[bal_idx] if bal_idx < len(row) else 0)

        cheque_raw = row[col["Chq./Ref.no."]] if col["Chq./Ref.no."] < len(row) else None

        seq += 1
        record = {
            "kind": "statement",
            "row_number": excel_row,
            "statement_no": seq,
            "transaction_id": "",
            "value_date": value_date,
            "posted_date": value_date,  # Ujjivan has no separate posted date
            "cheque_no": _clean_optional_text(cheque_raw),
            "description": description,
            "direction": direction,
            "amount": amount,
            "available_balance": available_balance,
            "refs": extract_ref_from_description(description),
            "is_fd": False,
            "match_state": "unmatched",
            "matched": False,
        }
        record["row_hash"] = hash_bank_stmt_row(
            account_id,
            date_to_iso(value_date) or "",
            str(amount),
            direction,
            description,
        )
        transactions.append(record)
        closing_balance = available_balance

    return {
        "sheet_name": sheet_title,
        "transactions": transactions,
        "count": len(transactions),
        "closing_balance": closing_balance,
        "period_start": transactions[0]["value_date"] if transactions else None,
        "period_end": transactions[-1]["value_date"] if transactions else None,
    }


# ===================================================================
# HDFC parsing
# ===================================================================

def _parse_hdfc(
    rows: list[tuple[Any, ...]],
    header_row_index: int,
    account_id: str,
    sheet_title: str,
) -> dict[str, Any]:
    """Parse HDFC-format statement rows.

    HDFC uses separate Credit / Debit columns, has Reference No. and 
    optional Cheque No./Instrument No. columns. Dates may be in various formats.
    """

    data_start = header_row_index + 1
    headers = [str(v).strip() if v is not None else "" for v in rows[header_row_index]]
    col = {h: i for i, h in enumerate(headers)}

    # Handle alternative column names for Cheque No.
    cheque_col_name = None
    for candidate in ["Cheque No./Instrument No.", "Cheque No.", "Instrument No."]:
        if candidate in col:
            cheque_col_name = candidate
            break

    transactions: list[dict[str, Any]] = []
    closing_balance = normalise_amount(0)
    seq = 0

    for excel_row, row in enumerate(rows[data_start:], start=data_start + 1):
        # Skip rows without a parseable date (footer / summary rows).
        date_idx = col.get("Date", 0)
        raw_date = row[date_idx] if date_idx < len(row) else None
        value_date = normalise_date(raw_date)
        if value_date is None:
            continue

        description_idx = col.get("Description", 1)
        description = str(row[description_idx] if description_idx < len(row) else "").strip()
        if not description:
            continue

        # Get Credit/Debit amounts
        credit_idx = col.get("Credit", -1)
        debit_idx = col.get("Debit", -1)
        credit = normalise_amount(row[credit_idx] if 0 <= credit_idx < len(row) else 0)
        debit = normalise_amount(row[debit_idx] if 0 <= debit_idx < len(row) else 0)

        if credit > 0:
            direction = "IN"
            amount = credit
        elif debit > 0:
            direction = "OUT"
            amount = debit
        else:
            continue  # skip zero-amount rows

        # Balance
        balance_idx = col.get("Balance", -1)
        available_balance = normalise_amount(row[balance_idx] if 0 <= balance_idx < len(row) else 0)

        # Reference No. as transaction_id
        ref_idx = col.get("Reference No.", -1)
        ref_no = _clean_optional_text(row[ref_idx] if 0 <= ref_idx < len(row) else None) or ""

        # Cheque No.
        cheque_raw = None
        if cheque_col_name and cheque_col_name in col:
            cheque_idx = col[cheque_col_name]
            cheque_raw = row[cheque_idx] if cheque_idx < len(row) else None

        # Value Date (HDFC has separate Date and Value Date columns)
        value_date_col_idx = col.get("Value Date", -1)
        if value_date_col_idx >= 0 and value_date_col_idx < len(row):
            vd = normalise_date(row[value_date_col_idx])
            if vd:
                value_date = vd

        seq += 1
        record = {
            "kind": "statement",
            "row_number": excel_row,
            "statement_no": seq,
            "transaction_id": ref_no,
            "value_date": value_date,
            "posted_date": value_date,  # HDFC doesn't have separate posted date
            "cheque_no": _clean_optional_text(cheque_raw),
            "description": description,
            "direction": direction,
            "amount": amount,
            "available_balance": available_balance,
            "refs": extract_ref_from_description(description),
            "is_fd": "REDEEM" in description.upper() or "FD" in description.upper(),
            "match_state": "unmatched",
            "matched": False,
        }
        record["row_hash"] = hash_bank_stmt_row(
            account_id,
            date_to_iso(value_date) or "",
            str(amount),
            direction,
            description,
        )
        transactions.append(record)
        closing_balance = available_balance

    return {
        "sheet_name": sheet_title,
        "transactions": transactions,
        "count": len(transactions),
        "closing_balance": closing_balance,
        "period_start": transactions[0]["value_date"] if transactions else None,
        "period_end": transactions[-1]["value_date"] if transactions else None,
    }


# ===================================================================
# HDFC Bank (Actual format) parsing
# ===================================================================

def _parse_hdfc_actual(
    rows: list[tuple[Any, ...]],
    header_row_index: int,
    account_id: str,
    sheet_title: str,
) -> dict[str, Any]:
    """Parse HDFC Bank actual format (Withdrawal Amt./Deposit Amt. columns)."""

    data_start = header_row_index + 1
    headers = [str(v).strip() if v is not None else "" for v in rows[header_row_index]]
    col = {h: i for i, h in enumerate(headers)}

    transactions: list[dict[str, Any]] = []
    closing_balance = normalise_amount(0)
    seq = 0

    for excel_row, row in enumerate(rows[data_start:], start=data_start + 1):
        # Skip separator rows with asterisks
        date_idx = col.get("Date", 0)
        raw_date = row[date_idx] if date_idx < len(row) else None
        if raw_date and str(raw_date).startswith("*"):
            continue
        value_date = normalise_date(raw_date)
        if value_date is None:
            continue

        narration_idx = col.get("Narration", 1)
        description = str(row[narration_idx] if narration_idx < len(row) else "").strip()
        if not description or description.startswith("*"):
            continue

        # Get Withdrawal/Deposit amounts
        withdrawal_idx = col.get("Withdrawal Amt.", -1)
        deposit_idx = col.get("Deposit Amt.", -1)
        withdrawal = normalise_amount(row[withdrawal_idx] if 0 <= withdrawal_idx < len(row) else 0)
        deposit = normalise_amount(row[deposit_idx] if 0 <= deposit_idx < len(row) else 0)

        if deposit > 0:
            direction = "IN"
            amount = deposit
        elif withdrawal > 0:
            direction = "OUT"
            amount = withdrawal
        else:
            continue

        # Balance
        balance_idx = col.get("Closing Balance", -1)
        available_balance = normalise_amount(row[balance_idx] if 0 <= balance_idx < len(row) else 0)

        # Cheque/Ref No.
        ref_idx = col.get("Chq./Ref.No.", -1)
        ref_no = _clean_optional_text(row[ref_idx] if 0 <= ref_idx < len(row) else None) or ""

        # Value Dt
        value_dt_idx = col.get("Value Dt", -1)
        if value_dt_idx >= 0 and value_dt_idx < len(row):
            vd = normalise_date(row[value_dt_idx])
            if vd:
                value_date = vd

        seq += 1
        record = {
            "kind": "statement",
            "row_number": excel_row,
            "statement_no": seq,
            "transaction_id": ref_no,
            "value_date": value_date,
            "posted_date": value_date,
            "cheque_no": _clean_optional_text(ref_no) if ref_no else None,
            "description": description,
            "direction": direction,
            "amount": amount,
            "available_balance": available_balance,
            "refs": extract_ref_from_description(description),
            "is_fd": "REDEEM" in description.upper() or "FD" in description.upper(),
            "match_state": "unmatched",
            "matched": False,
        }
        # Include Chq./Ref.No. in refs when it looks like a bank reference
        # (NEFT UTR, cheque number, etc.) — numeric strings of 6+ digits that
        # are not already captured from the narration.
        clean_ref = (ref_no or "").strip().lstrip("0") or ref_no
        if clean_ref and clean_ref.isdigit() and len(clean_ref) >= 6:
            if clean_ref not in record["refs"]:
                record["refs"].append(clean_ref)
        record["row_hash"] = hash_bank_stmt_row(
            account_id,
            date_to_iso(value_date) or "",
            str(amount),
            direction,
            description,
        )
        transactions.append(record)
        closing_balance = available_balance

    return {
        "sheet_name": sheet_title,
        "transactions": transactions,
        "count": len(transactions),
        "closing_balance": closing_balance,
        "period_start": transactions[0]["value_date"] if transactions else None,
        "period_end": transactions[-1]["value_date"] if transactions else None,
    }


# ===================================================================
# Axis Bank parsing
# ===================================================================

def _parse_axis(
    rows: list[tuple[Any, ...]],
    header_row_index: int,
    account_id: str,
    sheet_title: str,
) -> dict[str, Any]:
    """Parse Axis Bank format (Amount + Transaction Type CR/DR)."""

    data_start = header_row_index + 1
    headers = [str(v).strip() if v is not None else "" for v in rows[header_row_index]]
    col = {h: i for i, h in enumerate(headers)}

    transactions: list[dict[str, Any]] = []
    closing_balance = normalise_amount(0)
    seq = 0

    for excel_row, row in enumerate(rows[data_start:], start=data_start + 1):
        # Get Transaction Date
        date_idx = col.get("Transaction Date", -1)
        if date_idx < 0:
            continue
        raw_date = row[date_idx] if date_idx < len(row) else None
        value_date = normalise_date(raw_date)
        if value_date is None:
            continue

        # Description (Transaction Particulars)
        desc_idx = col.get("Transaction Particulars", -1)
        description = str(row[desc_idx] if 0 <= desc_idx < len(row) else "").strip()
        if not description:
            continue

        # Amount and Transaction Type
        amount_idx = col.get("Amount", -1)
        type_idx = col.get("Transaction Type", -1)
        amount = normalise_amount(row[amount_idx] if 0 <= amount_idx < len(row) else 0)
        tx_type = str(row[type_idx] if 0 <= type_idx < len(row) else "").strip().upper()

        if amount <= 0:
            continue

        direction = "IN" if tx_type == "CR" else "OUT"

        # Balance
        balance_idx = col.get("Balance", -1)
        available_balance = normalise_amount(row[balance_idx] if 0 <= balance_idx < len(row) else 0)

        # Cheque Number
        cheque_idx = col.get("Cheque Number", -1)
        cheque_no = _clean_optional_text(row[cheque_idx] if 0 <= cheque_idx < len(row) else None)

        # Value Date
        value_date_idx = col.get("Value Date", -1)
        if value_date_idx >= 0 and value_date_idx < len(row):
            vd = normalise_date(row[value_date_idx])
            if vd:
                value_date = vd

        seq += 1
        record = {
            "kind": "statement",
            "row_number": excel_row,
            "statement_no": seq,
            "transaction_id": "",
            "value_date": value_date,
            "posted_date": value_date,
            "cheque_no": cheque_no,
            "description": description,
            "direction": direction,
            "amount": amount,
            "available_balance": available_balance,
            "refs": extract_ref_from_description(description),
            "is_fd": False,
            "match_state": "unmatched",
            "matched": False,
        }
        record["row_hash"] = hash_bank_stmt_row(
            account_id,
            date_to_iso(value_date) or "",
            str(amount),
            direction,
            description,
        )
        transactions.append(record)
        closing_balance = available_balance

    return {
        "sheet_name": sheet_title,
        "transactions": transactions,
        "count": len(transactions),
        "closing_balance": closing_balance,
        "period_start": transactions[0]["value_date"] if transactions else None,
        "period_end": transactions[-1]["value_date"] if transactions else None,
    }


# ===================================================================
# IDBI Bank parsing
# ===================================================================

def _parse_idbi(
    rows: list[tuple[Any, ...]],
    header_row_index: int,
    account_id: str,
    sheet_title: str,
) -> dict[str, Any]:
    """Parse IDBI Bank format (GL. Date, Transaction Debit/Credit Amount).
    
    IDBI has a quirky format where the header spans two rows and columns may
    have empty cells that cause offsets in the data rows.
    """

    data_start = header_row_index + 2  # IDBI has split header rows
    headers = [str(v).strip() if v is not None else "" for v in rows[header_row_index]]
    col = {h: i for i, h in enumerate(headers)}

    transactions: list[dict[str, Any]] = []
    closing_balance = normalise_amount(0)
    seq = 0

    # Find the Debit Amount column index - IDBI puts empty cells in headers
    debit_hdr_idx = -1
    credit_hdr_idx = -1
    for i, h in enumerate(headers):
        if "Debit" in h and "Amount" in h:
            debit_hdr_idx = i
        if "Credit" in h and "Amount" in h:
            credit_hdr_idx = i

    # Pre-scan for an explicit "Closing Balance" row produced by IDBI reports.
    for row in rows[data_start:]:
        cell0 = str(row[0]).strip() if row and row[0] is not None else ""
        if cell0.lower().startswith("closing balance"):
            # Value is typically in the 3rd cell (index 2) as "39,43,590.94Cr"
            for cell in row[1:]:
                if cell is None:
                    continue
                raw = str(cell).strip()
                if raw == ":":
                    continue
                # Strip Cr/Dr suffix and commas
                cleaned = raw.replace("Cr", "").replace("Dr", "").replace(",", "").strip()
                parsed = normalise_amount(cleaned)
                if parsed > 0:
                    closing_balance = parsed
                    break
            break

    for excel_row, row in enumerate(rows[data_start:], start=data_start + 1):
        # Get GL Date
        date_idx = col.get("GL.", 0)
        raw_date = row[date_idx] if date_idx < len(row) else None
        value_date = normalise_date(raw_date)
        if value_date is None:
            continue

        # Description (Particulars)
        desc_idx = col.get("Particulars", -1)
        description = str(row[desc_idx] if 0 <= desc_idx < len(row) else "").strip()
        if not description:
            continue

        # IDBI format quirk: the Debit Amount header is at index 5, but the value
        # may be at index 6 (one cell shifted). Try both positions.
        debit = 0
        credit = 0
        
        if debit_hdr_idx >= 0:
            # Try the header position first
            debit = normalise_amount(row[debit_hdr_idx] if debit_hdr_idx < len(row) else 0)
            # If empty, try the next cell (IDBI offset quirk)
            if debit == 0 and debit_hdr_idx + 1 < len(row):
                debit = normalise_amount(row[debit_hdr_idx + 1])
        
        if credit_hdr_idx >= 0:
            credit = normalise_amount(row[credit_hdr_idx] if credit_hdr_idx < len(row) else 0)
            # If empty, try the next cell
            if credit == 0 and credit_hdr_idx + 1 < len(row):
                credit = normalise_amount(row[credit_hdr_idx + 1])

        if debit > 0:
            direction = "OUT"
            amount = debit
        elif credit > 0:
            direction = "IN"
            amount = credit
        else:
            continue

        # Balance - look for it after the credit column
        balance_idx = col.get("Balance", -1)
        if balance_idx < 0 and credit_hdr_idx >= 0:
            balance_idx = credit_hdr_idx + 1
        
        # IDBI balance may have "Cr" or "Dr" suffix, handle with normalise_amount
        raw_balance = row[balance_idx] if 0 <= balance_idx < len(row) else 0
        # IDBI offset quirk: balance value may be shifted by one column,
        # same as the debit/credit columns.
        if (not raw_balance or raw_balance is None) and 0 <= balance_idx + 1 < len(row):
            raw_balance = row[balance_idx + 1]
        if raw_balance and isinstance(raw_balance, str):
            # Remove Cr/Dr suffix and commas
            raw_balance = raw_balance.replace("Cr", "").replace("Dr", "").replace(",", "").strip()
        available_balance = normalise_amount(raw_balance)

        # Tran Id
        tran_id_idx = col.get("Tran Id", -1)
        tran_id = _clean_optional_text(row[tran_id_idx] if 0 <= tran_id_idx < len(row) else None) or ""

        # Instrument Number
        instr_idx = col.get("Instrmnt Number", -1)
        cheque_no = _clean_optional_text(row[instr_idx] if 0 <= instr_idx < len(row) else None)

        seq += 1
        record = {
            "kind": "statement",
            "row_number": excel_row,
            "statement_no": seq,
            "transaction_id": tran_id,
            "value_date": value_date,
            "posted_date": value_date,
            "cheque_no": cheque_no,
            "description": description,
            "direction": direction,
            "amount": amount,
            "available_balance": available_balance,
            "refs": extract_ref_from_description(description),
            "is_fd": False,
            "match_state": "unmatched",
            "matched": False,
        }
        record["row_hash"] = hash_bank_stmt_row(
            account_id,
            date_to_iso(value_date) or "",
            str(amount),
            direction,
            description,
        )
        transactions.append(record)
        if available_balance > 0:
            closing_balance = available_balance

    return {
        "sheet_name": sheet_title,
        "transactions": transactions,
        "count": len(transactions),
        "closing_balance": closing_balance,
        "period_start": transactions[0]["value_date"] if transactions else None,
        "period_end": transactions[-1]["value_date"] if transactions else None,
    }


# ===================================================================
# Karur Vysya (KV) Bank parsing
# ===================================================================

def _parse_kv(
    rows: list[tuple[Any, ...]],
    header_row_index: int,
    account_id: str,
    sheet_title: str,
) -> dict[str, Any]:
    """Parse Karur Vysya Bank format (Debit/Credit columns)."""

    data_start = header_row_index + 1
    headers = [str(v).strip() if v is not None else "" for v in rows[header_row_index]]
    col = {h: i for i, h in enumerate(headers)}

    transactions: list[dict[str, Any]] = []
    closing_balance = normalise_amount(0)
    seq = 0

    for excel_row, row in enumerate(rows[data_start:], start=data_start + 1):
        # Get Transaction Date
        date_idx = col.get("Transaction Date", 0)
        raw_date = row[date_idx] if date_idx < len(row) else None
        value_date = normalise_date(raw_date)
        if value_date is None:
            continue

        # Description
        desc_idx = col.get("Description", -1)
        description = str(row[desc_idx] if 0 <= desc_idx < len(row) else "").strip()
        if not description:
            continue

        # Debit/Credit amounts
        debit_idx = col.get("Debit", -1)
        credit_idx = col.get("Credit", -1)
        debit = normalise_amount(row[debit_idx] if 0 <= debit_idx < len(row) else 0)
        credit = normalise_amount(row[credit_idx] if 0 <= credit_idx < len(row) else 0)

        if credit > 0:
            direction = "IN"
            amount = credit
        elif debit > 0:
            direction = "OUT"
            amount = debit
        elif credit < 0:
            direction = "OUT"
            amount = abs(credit)
        elif debit < 0:
            direction = "IN"
            amount = abs(debit)
        else:
            continue

        # Balance
        balance_idx = col.get("Balance", -1)
        available_balance = normalise_amount(row[balance_idx] if 0 <= balance_idx < len(row) else 0)

        # Cheque No.
        cheque_idx = col.get("Cheque No.", -1)
        cheque_no = _clean_optional_text(row[cheque_idx] if 0 <= cheque_idx < len(row) else None)

        # Value Date
        value_date_idx = col.get("Value Date", -1)
        if value_date_idx >= 0 and value_date_idx < len(row):
            vd = normalise_date(row[value_date_idx])
            if vd:
                value_date = vd

        seq += 1
        record = {
            "kind": "statement",
            "row_number": excel_row,
            "statement_no": seq,
            "transaction_id": "",
            "value_date": value_date,
            "posted_date": value_date,
            "cheque_no": cheque_no,
            "description": description,
            "direction": direction,
            "amount": amount,
            "available_balance": available_balance,
            "refs": extract_ref_from_description(description),
            "is_fd": "REDEEM" in description.upper() or "FD" in description.upper(),
            "match_state": "unmatched",
            "matched": False,
        }
        record["row_hash"] = hash_bank_stmt_row(
            account_id,
            date_to_iso(value_date) or "",
            str(amount),
            direction,
            description,
        )
        transactions.append(record)
        closing_balance = available_balance

    return {
        "sheet_name": sheet_title,
        "transactions": transactions,
        "count": len(transactions),
        "closing_balance": closing_balance,
        "period_start": transactions[0]["value_date"] if transactions else None,
        "period_end": transactions[-1]["value_date"] if transactions else None,
    }


# ===================================================================
# SBI Bank parsing
# ===================================================================

def _parse_sbi(
    rows: list[tuple[Any, ...]],
    header_row_index: int,
    account_id: str,
    sheet_title: str,
) -> dict[str, Any]:
    """Parse SBI Bank format (Txn Date, Debit/Credit columns)."""

    data_start = header_row_index + 1
    headers = [str(v).strip() if v is not None else "" for v in rows[header_row_index]]
    col = {h: i for i, h in enumerate(headers)}

    transactions: list[dict[str, Any]] = []
    closing_balance = normalise_amount(0)
    seq = 0

    for excel_row, row in enumerate(rows[data_start:], start=data_start + 1):
        # Get Txn Date
        date_idx = col.get("Txn Date", 0)
        raw_date = row[date_idx] if date_idx < len(row) else None
        value_date = normalise_date(raw_date)
        if value_date is None:
            continue

        # Description
        desc_idx = col.get("Description", -1)
        description = str(row[desc_idx] if 0 <= desc_idx < len(row) else "").strip()
        if not description:
            continue

        # Debit/Credit amounts
        debit_idx = col.get("Debit", -1)
        credit_idx = col.get("Credit", -1)
        debit = normalise_amount(row[debit_idx] if 0 <= debit_idx < len(row) else 0)
        credit = normalise_amount(row[credit_idx] if 0 <= credit_idx < len(row) else 0)

        if credit > 0:
            direction = "IN"
            amount = credit
        elif debit > 0:
            direction = "OUT"
            amount = debit
        else:
            continue

        # Balance
        balance_idx = col.get("Balance", -1)
        available_balance = normalise_amount(row[balance_idx] if 0 <= balance_idx < len(row) else 0)

        # Ref No./Cheque No.
        ref_idx = col.get("Ref No./Cheque No.", -1)
        ref_no = _clean_optional_text(row[ref_idx] if 0 <= ref_idx < len(row) else None)

        # Value Date
        value_date_idx = col.get("Value Date", -1)
        if value_date_idx >= 0 and value_date_idx < len(row):
            vd = normalise_date(row[value_date_idx])
            if vd:
                value_date = vd

        seq += 1
        record = {
            "kind": "statement",
            "row_number": excel_row,
            "statement_no": seq,
            "transaction_id": ref_no or "",
            "value_date": value_date,
            "posted_date": value_date,
            "cheque_no": ref_no,
            "description": description,
            "direction": direction,
            "amount": amount,
            "available_balance": available_balance,
            "refs": extract_ref_from_description(description),
            "is_fd": False,
            "match_state": "unmatched",
            "matched": False,
        }
        record["row_hash"] = hash_bank_stmt_row(
            account_id,
            date_to_iso(value_date) or "",
            str(amount),
            direction,
            description,
        )
        transactions.append(record)
        closing_balance = available_balance

    return {
        "sheet_name": sheet_title,
        "transactions": transactions,
        "count": len(transactions),
        "closing_balance": closing_balance,
        "period_start": transactions[0]["value_date"] if transactions else None,
        "period_end": transactions[-1]["value_date"] if transactions else None,
    }


def _select_sheet(workbook: openpyxl.Workbook, explicit_name: str | None) -> openpyxl.worksheet.worksheet.Worksheet:
    """Select the statement sheet by explicit name or the expected naming pattern."""

    if explicit_name:
        if explicit_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{explicit_name}' not found in {workbook.sheetnames}")
        return workbook[explicit_name]

    for candidate in workbook.sheetnames:
        if SHEET_PATTERN.search(candidate):
            return workbook[candidate]

    for candidate in workbook.worksheets:
        if _sheet_has_statement_schema(candidate):
            return candidate

    if len(workbook.sheetnames) == 1:
        return workbook.active

    raise ValueError(f"No bank statement sheet found in workbook: {workbook.sheetnames}")


def _sheet_has_statement_schema(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
    """Detect the statement sheet from its headers even if the tab is generic."""

    max_scan = max(
        UJJIVAN_HEADER_SCAN_LIMIT,
        HDFC_ACTUAL_HEADER_SCAN_LIMIT,
        IDBI_HEADER_SCAN_LIMIT,
        SBI_HEADER_SCAN_LIMIT,
    )
    rows = list(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(max_scan, worksheet.max_row),
            values_only=True,
        )
    )
    return (
        _find_header_row(rows, ICICI_REQUIRED_COLUMNS, STATEMENT_HEADER_SCAN_LIMIT) is not None
        or _find_header_row(rows, UJJIVAN_REQUIRED_COLUMNS, UJJIVAN_HEADER_SCAN_LIMIT) is not None
        or _find_header_row(rows, HDFC_REQUIRED_COLUMNS, HDFC_HEADER_SCAN_LIMIT) is not None
        or _find_header_row(rows, HDFC_ACTUAL_REQUIRED_COLUMNS, HDFC_ACTUAL_HEADER_SCAN_LIMIT) is not None
        or _find_header_row(rows, AXIS_REQUIRED_COLUMNS, AXIS_HEADER_SCAN_LIMIT) is not None
        or _find_header_row(rows, IDBI_REQUIRED_COLUMNS, IDBI_HEADER_SCAN_LIMIT) is not None
        or _find_header_row(rows, KV_REQUIRED_COLUMNS, KV_HEADER_SCAN_LIMIT) is not None
        or _find_header_row(rows, SBI_REQUIRED_COLUMNS, SBI_HEADER_SCAN_LIMIT) is not None
    )


def _find_header_row(
    rows: list[tuple[Any, ...]],
    required_columns: list[str],
    scan_limit: int,
) -> int | None:
    """Return the zero-based row index whose cells contain all *required_columns*."""

    limit = min(scan_limit, len(rows))
    for row_index in range(limit):
        headers = {str(value).strip() for value in rows[row_index] if value is not None}
        if all(column in headers for column in required_columns):
            return row_index
    return None


# Keep the old name so any external callers still work.
def _find_statement_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    return _find_header_row(rows, ICICI_REQUIRED_COLUMNS, STATEMENT_HEADER_SCAN_LIMIT)


def _clean_optional_text(raw: Any) -> str | None:
    """Return a cleaned text value or ``None`` for blank placeholders."""

    text = str(raw).strip() if raw is not None else ""
    if not text or text == "-":
        return None
    return text
