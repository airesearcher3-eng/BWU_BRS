"""Parser for the previous month's BRS carry-forward sheet."""

from __future__ import annotations

from pathlib import Path
from typing import Any
import re

import openpyxl

from engine.normaliser import normalise_amount, normalise_date
from engine.reference_extractor import extract_ref_from_description


def _extract_reconciled_balance(worksheet) -> "Decimal | None":
    """Scan the BRS sheet for the 'Reconcile Balance' row and return column-G value."""
    from decimal import Decimal
    for row in worksheet.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        cell_a = str(row[0]).strip().lower()
        if "reconcile balance" in cell_a or "reconciled balance" in cell_a:
            val = row[6] if len(row) > 6 else None
            if val is not None:
                try:
                    return normalise_amount(val)
                except Exception:
                    pass
    return None


def _extract_bank_book_balance(worksheet) -> "Decimal | None":
    """Scan the BRS sheet for the opening 'Balance as per Bank Book' row (column G)."""
    for row in worksheet.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        cell_a = str(row[0]).strip().lower()
        if "balance as per bank book" in cell_a and "reconcile" not in cell_a:
            val = row[6] if len(row) > 6 else None
            if val is not None:
                try:
                    return normalise_amount(val)
                except Exception:
                    pass
    return None


SHEET_PATTERN = re.compile(r"^BRS", re.IGNORECASE)
SECTION_LABELS = {
    "add_cheque_issued": "add: cheque issued but not debited by bank",
    "add_bank_credit": "add: amount credited by bank but not entered in bank book",
    "less_cheque_deposit": "less: cheque deposited but not credited by bank",
    "less_bank_debit": "less: amount debited by bank but not entered in bank book",
}

# Alternative phrasings found in manually-prepared BRS files.
_SECTION_ALT_LABELS: dict[str, list[str]] = {
    "add_cheque_issued": [
        "add: cheques issued but not debited by bank",
    ],
    "less_cheque_deposit": [
        "less: deposited into bank but not credited",
    ],
}


def parse_previous_brs(
    filepath: str | Path,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Parse all carry-forward items, preserving both pending and resolved rows."""

    workbook = openpyxl.load_workbook(filepath, data_only=True)
    worksheet = _select_sheet(workbook, sheet_name)

    items: list[dict[str, Any]] = []
    current_section: str | None = None
    # Auto-detect column layout from the first section header row.
    # Generated BRS: Date(A) | Remarks(B) | Chq(C) | Amount(D) | Cleared(E)
    # Manual BRS:    Remarks(A) | Chq(B) | Date(C) | Amount(D) | Cleared(E)
    col_layout: str | None = None  # "generated" or "manual"

    for excel_row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        cell_a = str(row[0] or "").strip()
        section = _detect_section(cell_a)
        if section:
            current_section = section
            col_layout = None  # Reset layout for each section to detect from header
            # Check if column headers are on the SAME row as the section label
            # (common in manually-prepared BRS files).
            if len(row) >= 5:
                col_b = str(row[1] or "").strip().lower()
                col_c = str(row[2] or "").strip().lower()
                if "chq" in col_b and "date" in col_c:
                    col_layout = "manual"
                elif "date" in col_c and "cheque" in col_b:
                    col_layout = "manual"
                elif "date" in col_b:
                    col_layout = "generated"
            continue

        # Detect layout from header row (row after section label)
        if current_section and col_layout is None and len(row) >= 5:
            col_a = str(row[0] or "").strip().lower()
            col_b = str(row[1] or "").strip().lower()
            col_c = str(row[2] or "").strip().lower()
            # Bandhan/Manual layout: Chq(A) | Date(B) | Amount(C) | Remarks(D) | Cleared(E)
            if "chq" in col_a and "date" in col_b and "amount" in col_c:
                col_layout = "bandhan"
                continue
            # Manual layout: Remarks(A) | Chq(B) | Date(C) | Amount(D) | Cleared(E)
            elif "date" in col_c and "cheque" in col_b:
                col_layout = "manual"
                continue
            # Generated layout: Date(A) | Remarks(B) | Chq(C) | Amount(D) | Cleared(E)
            elif "date" in col_a or "date" in col_b:
                col_layout = "generated"
                continue

        if not current_section or len(row) < 5:
            continue

        # Try column layouts based on detected format.
        if col_layout == "bandhan":
            # Bandhan: Chq(A) | Date(B) | Amount(C) | Remarks(D) | Cleared(E)
            original_date = normalise_date(row[1])
            amount = normalise_amount(row[2])
            remarks_raw = row[3]
            cheque_raw = row[0]
        elif col_layout == "manual":
            # Manual: Remarks(A) | Chq(B) | Date(C) | Amount(D) | Cleared(E)
            original_date = normalise_date(row[2])
            amount = normalise_amount(row[3])
            remarks_raw = row[0]
            cheque_raw = row[1]
        else:
            # Generated: Date(A) | Remarks(B) | Chq(C) | Amount(D) | Cleared(E)
            original_date = normalise_date(row[0])
            amount = normalise_amount(row[3])
            remarks_raw = row[1]
            cheque_raw = row[2]

        if original_date is None or amount <= 0:
            # Fallback: try bandhan layout if no date found with default layout
            if col_layout is None:
                alt_date = normalise_date(row[1])
                alt_amount = normalise_amount(row[2])
                if alt_date and alt_amount > 0:
                    original_date = alt_date
                    amount = alt_amount
                    remarks_raw = row[3]
                    cheque_raw = row[0]
                    col_layout = "bandhan"
                else:
                    continue
            else:
                continue

        item = {
            "row_number": excel_row_number,
            "section": current_section,
            "original_date": original_date,
            "remarks": str(remarks_raw or "").strip(),
            "cheque_no": _clean_optional_text(cheque_raw),
            "amount": amount,
            "cleared_on": normalise_date(row[4]),
        }
        item["direction"] = _direction_for_section(current_section)
        item["is_pending"] = item["cleared_on"] is None
        item["refs"] = extract_ref_from_description(item["remarks"])
        items.append(item)

    reconciled_balance = _extract_reconciled_balance(worksheet)
    bank_book_balance = _extract_bank_book_balance(worksheet)

    workbook.close()

    pending = [item for item in items if item["is_pending"]]
    resolved = [item for item in items if not item["is_pending"]]
    return {
        "sheet_name": worksheet.title,
        "items": items,
        "pending_items": pending,
        "resolved_items": resolved,
        "count": len(items),
        # The reconciled balance from the previous period is what the
        # current period's bank-book opening balance should equal.
        "reconciled_balance": reconciled_balance,
        # The bank-book balance as stated in the previous BRS header.
        "bank_book_balance": bank_book_balance,
    }


def _select_sheet(workbook: openpyxl.Workbook, explicit_name: str | None) -> openpyxl.worksheet.worksheet.Worksheet:
    """Select the BRS sheet by explicit name or the first matching title."""

    if explicit_name:
        if explicit_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{explicit_name}' not found in {workbook.sheetnames}")
        return workbook[explicit_name]

    for candidate in workbook.sheetnames:
        if SHEET_PATTERN.search(candidate):
            return workbook[candidate]

    for candidate in workbook.worksheets:
        if _sheet_has_brs_markers(candidate):
            return candidate

    if len(workbook.sheetnames) == 1:
        return workbook.active

    raise ValueError(f"No BRS sheet found in workbook: {workbook.sheetnames}")


def _sheet_has_brs_markers(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
    """Detect a BRS report sheet from its section labels and title lines."""

    first_col_values = []
    for row in range(1, min(80, worksheet.max_row) + 1):
        value = worksheet.cell(row=row, column=1).value
        if value is not None:
            first_col_values.append(str(value).strip().lower())

    return (
        any("bank reconcillation statement" in value for value in first_col_values)
        and any(label in value for value in first_col_values for label in SECTION_LABELS.values())
    )


def _detect_section(text: str) -> str | None:
    """Return the logical BRS section for a heading row."""

    lowered = text.lower()
    for section, label in SECTION_LABELS.items():
        if label in lowered:
            return section
    # Check alternative phrasings.
    for section, alt_labels in _SECTION_ALT_LABELS.items():
        for alt in alt_labels:
            if alt in lowered:
                return section
    return None


def _direction_for_section(section: str) -> str:
    """Infer whether the carry-forward item is an ``IN`` or ``OUT`` item."""

    if section in {"add_cheque_issued", "less_bank_debit"}:
        return "OUT"
    return "IN"


def _clean_optional_text(raw: Any) -> str | None:
    """Return a cleaned text value or ``None`` when the cell is blank."""

    text = str(raw).strip() if raw is not None else ""
    return text or None
