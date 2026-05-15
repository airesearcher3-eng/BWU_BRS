"""FastAPI routes for running and downloading reconciliations."""
from __future__ import annotations

import asyncio
import functools
import hashlib
import io
import json
import os
import tempfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel

from engine.reconciliation import reconcile_workbooks, serialise_result
from engine.normaliser import date_to_iso, decimal_to_float
from engine.brs_output import generate_brs_excel
from routes.auth import require_role
from models.database import (
    get_bank_account,
    get_connection,
    get_match_report,
    get_run,
    insert_audit_log,
    insert_exception,
    insert_match,
    insert_run,
    insert_transaction,
    update_run,
    update_transaction_status,
)

router = APIRouter(prefix="/api/reconciliation", tags=["Reconciliation"])
BASE_DIR = Path(__file__).resolve().parents[1]

EXCEPTION_SLA_DAYS = {
    "unknown_dr": 1,
    "amount_mismatch": 1,
    "gib_unmatched": 3,
    "unknown_cr": 3,
    "stale_carry_forward": 0,
    "timing_difference": 0,
}


class ReconciliationRequest(BaseModel):
    period_start: Optional[str] = None
    period_end: Optional[str] = None
    bank_statement_path: str
    bank_book_path: str
    previous_brs_path: Optional[str] = None
    previous_brs_sheet: Optional[str] = None
    portal_data_path: Optional[str] = None
    bank_account_id: Optional[int] = None


async def _run_reconciliation_task(
    run_id: int,
    req_dict: dict,
    bank_account: dict | None,
) -> None:
    """Background task: runs the full reconciliation pipeline and updates the DB."""
    output_dir = BASE_DIR / "output"
    output_path = output_dir / f"BRS_run_{run_id}.xlsx"
    try:
        result = await reconcile_workbooks(
            statement_path=req_dict["bank_statement_path"],
            bank_book_path=req_dict["bank_book_path"],
            previous_brs_path=req_dict.get("previous_brs_path"),
            previous_brs_sheet=req_dict.get("previous_brs_sheet"),
            portal_data_path=req_dict.get("portal_data_path"),
            output_path=output_path,
            bank_account=bank_account,
            use_rag=True,
        )
        summary = serialise_result(result)

        async with get_connection() as conn:
            await _persist_run_artifacts(conn, run_id, result)
            await update_run(
                conn,
                run_id,
                status="completed",
                period_start=(req_dict.get("period_start") or result["statement"]["period_start"].isoformat()),
                period_end=(req_dict.get("period_end") or result["statement"]["period_end"].isoformat()),
                bank_book_balance=summary["totals"]["bank_book_balance"],
                bank_statement_balance=summary["totals"]["bank_statement_balance"],
                total_bank_stmt_entries=summary["statement_count"],
                total_bank_book_entries=summary["bank_book_count"],
                pass1_matches=result["matching"]["pass_counts"].get(1, 0),
                pass2_matches=result["matching"]["pass_counts"].get(2, 0),
                pass3_matches=result["matching"]["pass_counts"].get(3, 0),
                pass4_matches=result["matching"]["pass_counts"].get(4, 0),
                total_matched=summary["statement_count"] - len(result["matching"]["unmatched_statement"]),
                total_unmatched=len(result["matching"]["unmatched_statement"]) + len(result["matching"]["unmatched_book"]),
                total_pending=len(result["exceptions"]),
                brs_output_path=str(output_path),
                completed_at=datetime.now(),
            )
            await insert_audit_log(
                conn, "run_completed",
                entity_type="run", entity_id=run_id,
                details=summary,
            )
    except Exception as exc:
        try:
            async with get_connection() as conn:
                await update_run(conn, run_id, status="failed")
                await insert_audit_log(
                    conn, "run_failed",
                    entity_type="run", entity_id=run_id,
                    details={"error": str(exc)},
                )
        except Exception:
            pass


@router.post("/run")
async def start_reconciliation(req: ReconciliationRequest, background_tasks: BackgroundTasks):
    """Start the reconciliation pipeline. Returns immediately with run_id; poll GET /run/{id} for status."""

    if not os.path.exists(req.bank_statement_path):
        raise HTTPException(400, f"Bank statement file not found: {req.bank_statement_path}")
    if not os.path.exists(req.bank_book_path):
        raise HTTPException(400, f"Bank book file not found: {req.bank_book_path}")

    bank_account = None
    if req.bank_account_id:
        async with get_connection() as conn:
            bank_account = await get_bank_account(conn, req.bank_account_id)
        if not bank_account:
            raise HTTPException(400, f"Bank account not found: {req.bank_account_id}")

    async with get_connection() as conn:
        run_id = await insert_run(
            conn,
            req.period_start or datetime.now().date().isoformat(),
            req.period_end or datetime.now().date().isoformat(),
            req.bank_statement_path,
            req.bank_book_path,
            req.previous_brs_path,
            bank_account_id=req.bank_account_id,
        )
        await insert_audit_log(conn, "run_started", entity_type="run", entity_id=run_id)

    background_tasks.add_task(
        _run_reconciliation_task,
        run_id,
        req.model_dump(),
        bank_account,
    )

    return {"run_id": run_id, "status": "running"}


@router.get("/runs")
async def list_runs():
    async with get_connection() as conn:
        rows = await conn.fetch("SELECT * FROM runs ORDER BY created_at DESC LIMIT 50")
    return [dict(r) for r in rows]


@router.delete("/run/{run_id}")
async def delete_run(run_id: int, _user: dict = Depends(require_role("system_admin", "finance_controller"))):
    """Delete a reconciliation run and all associated data (cascaded by DB)."""
    async with get_connection() as conn:
        run = await get_run(conn, run_id)
        if not run:
            raise HTTPException(404, "Run not found")
        # Remove output file from disk if it exists
        output_path = run.get("brs_output_path")
        if output_path and os.path.exists(output_path):
            try:
                os.remove(output_path)
            except OSError:
                pass
        await conn.execute("DELETE FROM runs WHERE id = $1", run_id)
    return {"deleted": run_id}


@router.get("/run/{run_id}")
async def get_run_details(run_id: int):
    async with get_connection() as conn:
        run = await get_run(conn, run_id)
        if not run:
            raise HTTPException(404, "Run not found")

        section_totals = await _reconstruct_section_totals(conn, run_id)
        if section_totals and run.get("bank_book_balance") is not None:
            bb = run["bank_book_balance"]
            reconciled = (
                bb
                + section_totals["add_cheque_issued"]
                + section_totals["add_bank_credit"]
                - section_totals["less_cheque_deposit"]
                - section_totals["less_bank_debit"]
            )
            run["totals"] = {
                "bank_book_balance": bb,
                "bank_statement_balance": run.get("bank_statement_balance", 0),
                "add_cheque_issued": section_totals["add_cheque_issued"],
                "add_bank_credit": section_totals["add_bank_credit"],
                "less_cheque_deposit": section_totals["less_cheque_deposit"],
                "less_bank_debit": section_totals["less_bank_debit"],
                "reconciled_balance": reconciled,
                "difference": reconciled - (run.get("bank_statement_balance") or 0),
            }
            run["section_summary"] = section_totals.get("section_summary", {})

        approvals = await conn.fetch(
            "SELECT * FROM approvals WHERE run_id=$1 ORDER BY level", run_id
        )
        run["approvals"] = [dict(a) for a in approvals]

    return run


@router.get("/run/{run_id}/matches")
async def get_run_matches(run_id: int):
    async with get_connection() as conn:
        report = await get_match_report(conn, run_id)
        if not report:
            raise HTTPException(404, "Run not found")
    return report


@router.get("/run/{run_id}/matches/download")
def _strip_tz(val):
    """Remove tzinfo from datetime/date so openpyxl can write the cell."""
    if isinstance(val, datetime) and val.tzinfo is not None:
        return val.replace(tzinfo=None)
    return val


async def download_matches_excel(run_id: int):
    async with get_connection() as conn:
        report = await get_match_report(conn, run_id)
        if not report:
            raise HTTPException(404, "Run not found")

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws_summary.append([f"Matched Report — Run #{run_id}"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([])
    ws_summary.append(["Period", f"{report.get('period_start', '')} to {report.get('period_end', '')}"])
    ws_summary.append(["Completed At", _strip_tz(report.get("completed_at", ""))])
    ws_summary.append(["Total Match Groups", report["match_count"]])
    ws_summary.append(["Statement Entries Matched", report["statement_entry_count"]])
    ws_summary.append(["Book Entries Matched", report["bank_book_entry_count"]])
    ws_summary.append(["Total Matched Amount", report["total_matched_amount"]])
    ws_summary.append([])

    pass_counts = Counter(m["pass_number"] for m in report["matches"])
    type_counts = Counter(m["match_type"] for m in report["matches"])

    ws_summary.append(["Pass-wise Breakdown"])
    ws_summary[f"A{ws_summary.max_row}"].font = header_font
    for pn in sorted(pass_counts):
        ws_summary.append([f"  Pass {pn}", pass_counts[pn]])
    ws_summary.append([])
    ws_summary.append(["Match Type Breakdown"])
    ws_summary[f"A{ws_summary.max_row}"].font = header_font
    for mt, cnt in type_counts.most_common():
        ws_summary.append([f"  {mt}", cnt])

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    ws = wb.create_sheet("Matches")
    detail_headers = [
        "Match #", "Pass", "Match Type", "Confidence", "Matched Amount", "Notes",
        "Stmt Date", "Stmt Amount", "Stmt Direction", "Stmt Description", "Stmt Refs",
        "Book Date", "Book Amount", "Book Direction", "Book Narration",
        "Book Voucher Type", "Book Voucher No", "Book Refs",
    ]
    ws.append(detail_headers)
    for col_idx in range(1, len(detail_headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for idx, match in enumerate(report["matches"], 1):
        stmt_entries = match.get("statement_entries", [])
        book_entries = match.get("bank_book_entries", [])
        for row_offset in range(max(len(stmt_entries), len(book_entries), 1)):
            row_data = [idx, match["pass_number"], match["match_type"],
                        match["confidence"], match["matched_amount"], match["notes"]] \
                if row_offset == 0 else ["", "", "", "", "", ""]
            if row_offset < len(stmt_entries):
                s = stmt_entries[row_offset]
                refs = s.get("references_json") or []
                if isinstance(refs, str): refs = json.loads(refs)
                row_data += [_strip_tz(s.get("transaction_date", "")), s.get("amount", ""),
                             s.get("direction", ""),
                             s.get("description") or s.get("narration", ""),
                             ", ".join(str(r) for r in refs)]
            else:
                row_data += ["", "", "", "", ""]
            if row_offset < len(book_entries):
                b = book_entries[row_offset]
                brefs = b.get("references_json") or []
                if isinstance(brefs, str): brefs = json.loads(brefs)
                row_data += [_strip_tz(b.get("transaction_date", "")), b.get("amount", ""),
                             b.get("direction", ""),
                             b.get("narration") or b.get("description", ""),
                             b.get("voucher_type", ""), b.get("voucher_no", ""),
                             ", ".join(str(r) for r in brefs)]
            else:
                row_data += ["", "", "", "", "", "", ""]
            ws.append(row_data)

    col_widths = [8, 6, 22, 10, 15, 30, 12, 14, 6, 45, 20, 12, 14, 6, 45, 12, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Matched_Report_Run_{run_id}.xlsx"'},
    )


@router.get("/run/{run_id}/download")
async def download_brs(run_id: int):
    async with get_connection() as conn:
        run = await get_run(conn, run_id)
        if not run:
            raise HTTPException(404, "Run not found")

    output_path = run.get("brs_output_path")
    if output_path and os.path.exists(output_path):
        return FileResponse(
            output_path,
            filename=os.path.basename(output_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # File missing (e.g. volume wiped) — regenerate from DB
    async with get_connection() as conn:
        buffer = await _regenerate_brs_from_db(conn, run)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="BRS_run_{run_id}.xlsx"'},
    )


# ── Helpers ───────────────────────────────────────────────────────

async def _regenerate_brs_from_db(conn, run: dict) -> io.BytesIO:
    """Rebuild the BRS Excel entirely from DB records when the file is missing."""
    run_id = run["id"]

    # Unmatched bank-book rows → cheque-issued (OUT) / cheque-deposit (IN)
    book_rows = await conn.fetch(
        "SELECT transaction_date, narration, description, cheque_no, amount, direction "
        "FROM transactions WHERE run_id=$1 AND source='bank_book' AND match_status='unmatched' "
        "ORDER BY transaction_date", run_id,
    )
    # Unmatched statement rows → bank-credit (IN) / bank-debit (OUT)
    stmt_rows = await conn.fetch(
        "SELECT transaction_date, description, narration, cheque_no, amount, direction "
        "FROM transactions WHERE run_id=$1 AND source='bank_statement' AND match_status='unmatched' "
        "ORDER BY transaction_date", run_id,
    )
    # Carry-forward items
    cf_rows = await conn.fetch(
        "SELECT brs_section, original_date, remarks, cheque_no, amount, cleared_date "
        "FROM carry_forward WHERE run_id=$1 ORDER BY original_date", run_id,
    )

    sections: dict[str, list] = {
        "add_cheque_issued": [], "add_bank_credit": [],
        "less_cheque_deposit": [], "less_bank_debit": [],
    }

    def _parse_date(val):
        if val is None:
            return None
        if isinstance(val, date):
            return val
        try:
            return date.fromisoformat(str(val)[:10])
        except (ValueError, TypeError):
            return None

    for r in book_rows:
        key = "add_cheque_issued" if r["direction"] == "OUT" else "less_cheque_deposit"
        sections[key].append({"date": _parse_date(r["transaction_date"]),
                               "amount": float(r["amount"] or 0),
                               "remarks": r["narration"] or r["description"] or "",
                               "cheque_no": r["cheque_no"]})
    for r in stmt_rows:
        key = "add_bank_credit" if r["direction"] == "IN" else "less_bank_debit"
        sections[key].append({"date": _parse_date(r["transaction_date"]),
                               "amount": float(r["amount"] or 0),
                               "remarks": r["description"] or r["narration"] or "",
                               "cheque_no": r["cheque_no"]})
    for r in cf_rows:
        sec = r["brs_section"]
        if sec in sections:
            sections[sec].append({"date": _parse_date(r["original_date"]),
                                   "amount": float(r["amount"] or 0),
                                   "remarks": r["remarks"] or "", "cheque_no": r["cheque_no"],
                                   "cleared_on": _parse_date(r["cleared_date"])})

    bb = float(run.get("bank_book_balance") or 0)
    bs = float(run.get("bank_statement_balance") or 0)
    add_cheque   = sum(i["amount"] for i in sections["add_cheque_issued"])
    add_credit   = sum(i["amount"] for i in sections["add_bank_credit"])
    less_deposit = sum(i["amount"] for i in sections["less_cheque_deposit"])
    less_debit   = sum(i["amount"] for i in sections["less_bank_debit"])
    reconciled   = bb + add_cheque + add_credit - less_deposit - less_debit
    totals = {"bank_book_balance": bb, "bank_statement_balance": bs,
              "add_cheque_issued": add_cheque, "add_bank_credit": add_credit,
              "less_cheque_deposit": less_deposit, "less_bank_debit": less_debit,
              "reconciled_balance": reconciled, "difference": reconciled - bs}

    bank_account = None
    if run.get("bank_account_id"):
        bank_account = await get_bank_account(conn, run["bank_account_id"])

    as_on_date = run.get("period_end")
    if isinstance(as_on_date, str):
        as_on_date = date.fromisoformat(as_on_date)
    if not as_on_date:
        as_on_date = date.today()

    # generate_brs_excel writes to a path; use a temp file then stream back
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(
            None,
            functools.partial(
                generate_brs_excel, tmp_path,
                as_on_date=as_on_date, bank_book_balance=bb,
                bank_statement_balance=bs, sections=sections,
                totals=totals, bank_account=bank_account,
            ),
        )
        with open(tmp_path, "rb") as f:
            buf = io.BytesIO(f.read())
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
    buf.seek(0)
    return buf


async def _reconstruct_section_totals(conn, run_id: int) -> dict | None:
    book_unmatched = await conn.fetch(
        "SELECT direction, SUM(amount) AS total, COUNT(*) AS cnt "
        "FROM transactions WHERE run_id=$1 AND source='bank_book' AND match_status='unmatched' "
        "GROUP BY direction",
        run_id,
    )
    stmt_unmatched = await conn.fetch(
        "SELECT direction, SUM(amount) AS total, COUNT(*) AS cnt "
        "FROM transactions WHERE run_id=$1 AND source='bank_statement' AND match_status='unmatched' "
        "GROUP BY direction",
        run_id,
    )
    cf_rows = await conn.fetch(
        "SELECT brs_section, SUM(amount) AS total, COUNT(*) AS cnt "
        "FROM carry_forward WHERE run_id=$1 GROUP BY brs_section",
        run_id,
    )

    if not book_unmatched and not stmt_unmatched and not cf_rows:
        return None

    sections = {k: 0.0 for k in ("add_cheque_issued", "add_bank_credit",
                                   "less_cheque_deposit", "less_bank_debit")}
    counts = {k: 0 for k in sections}

    for row in book_unmatched:
        key = "add_cheque_issued" if row["direction"] == "OUT" else "less_cheque_deposit"
        sections[key] += float(row["total"] or 0)
        counts[key] += row["cnt"] or 0

    for row in stmt_unmatched:
        key = "add_bank_credit" if row["direction"] == "IN" else "less_bank_debit"
        sections[key] += float(row["total"] or 0)
        counts[key] += row["cnt"] or 0

    for row in cf_rows:
        sec = row["brs_section"]
        if sec in sections:
            sections[sec] += float(row["total"] or 0)
            counts[sec] += row["cnt"] or 0

    return {
        **sections,
        "section_summary": {s: {"count": counts[s], "total": sections[s]} for s in sections},
    }


async def _persist_run_artifacts(conn, run_id: int, result: dict) -> None:
    statement_rows = result["statement"]["transactions"]
    bank_book_rows = result["bank_book"]["transactions"]
    carry_forward_items = result["matching"]["pending_carry_forward_items"]
    matches = result["matching"]["matches"]

    stmt_pass_map, book_pass_map = _build_match_pass_maps(matches)
    statement_ids: dict[int, int] = {}
    bank_book_ids: dict[int, int] = {}

    for row in statement_rows:
        tid = await insert_transaction(
            conn, run_id, "bank_statement",
            date_to_iso(row["value_date"]) or "",
            decimal_to_float(row["amount"]),
            row["direction"],
            references=row.get("refs", []),
            description=row["description"],
            cheque_no=row.get("cheque_no"),
            transaction_id=row.get("transaction_id"),
            original_row=row["row_number"],
            sha256_hash=row["row_hash"],
        )
        statement_ids[row["row_number"]] = tid
        if row.get("matched"):
            await update_transaction_status(conn, tid, "matched",
                                             stmt_pass_map.get(row["row_number"]))

    for row in bank_book_rows:
        tid = await insert_transaction(
            conn, run_id, "bank_book",
            date_to_iso(row["voucher_date"]) or "",
            decimal_to_float(row["amount"]),
            row["direction"],
            references=row.get("refs", []),
            narration=row["narration"],
            description=row["particulars"],
            voucher_type=row.get("voucher_type"),
            voucher_no=row.get("voucher_no"),
            cheque_no=row.get("cheque_no"),
            original_row=row["row_number"],
            sha256_hash=row["row_hash"],
        )
        bank_book_ids[row["row_number"]] = tid
        if row.get("matched"):
            await update_transaction_status(conn, tid, "matched",
                                             book_pass_map.get(row["row_number"]))

    for item in carry_forward_items:
        tid = await insert_transaction(
            conn, run_id, "carry_forward",
            date_to_iso(item["original_date"]) or "",
            decimal_to_float(item["amount"]),
            item.get("direction", "IN"),
            references=item.get("refs", []),
            narration=item["remarks"],
            description=item["remarks"],
            cheque_no=item.get("cheque_no"),
            original_row=item["row_number"],
            sha256_hash=_carry_forward_hash(run_id, item),
        )
        await conn.execute(
            """INSERT INTO carry_forward
               (run_id, brs_section, original_date, remarks, cheque_no, amount, cleared_date)
               VALUES ($1,$2,$3,$4,$5,$6,$7)""",
            run_id,
            item["section"],
            date_to_iso(item["original_date"]) or "",
            item["remarks"],
            item.get("cheque_no"),
            decimal_to_float(item["amount"]),
            date_to_iso(item.get("cleared_on")),
        )

    for match in matches:
        pass_number = match.get("pass_number")
        if pass_number not in {1, 2, 3, 4, 5, 6}:
            continue
        stmt_ids = [statement_ids[rn] for rn in match.get("statement_rows", [])
                    if rn in statement_ids]
        book_ids = [bank_book_ids[rn] for rn in match.get("book_rows", [])
                    if rn in bank_book_ids]
        if stmt_ids or book_ids:
            await insert_match(
                conn, run_id, min(pass_number, 6), match["match_type"],
                stmt_ids, book_ids,
                decimal_to_float(match["amount"]),
                notes=match.get("notes"),
            )

    # Persist exceptions for unmatched items
    exceptions_data = result.get("exceptions", [])
    for exc in exceptions_data:
        source = exc.get("source", "statement")
        row_number = exc.get("row_number")
        txn_id = (statement_ids if source == "statement" else bank_book_ids).get(row_number)
        if not txn_id:
            continue
        exc_type = exc.get("exception_type", "unknown_dr")
        sla = EXCEPTION_SLA_DAYS.get(exc_type, 3)
        brs_section = exc.get("brs_section", "less_bank_debit")
        await insert_exception(conn, run_id, txn_id, exc_type, brs_section, sla)


def _build_match_pass_maps(matches: list) -> tuple[dict, dict]:
    stmt_map: dict[int, int] = {}
    book_map: dict[int, int] = {}
    for m in matches:
        pass_num = m.get("pass_number", 0)
        for rn in m.get("statement_rows", []):
            stmt_map.setdefault(rn, pass_num)
        for rn in m.get("book_rows", []):
            book_map.setdefault(rn, pass_num)
    return stmt_map, book_map


def _carry_forward_hash(run_id: int, item: dict) -> str:
    raw = f"cf|{run_id}|{item['row_number']}|{item['amount']}|{item['original_date']}"
    return hashlib.sha256(raw.encode()).hexdigest()
